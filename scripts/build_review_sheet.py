#!/usr/bin/env python3
"""Build a translation-review spreadsheet (Excel .xlsx) for play-through QA.

Joins the Japanese source (texts/ja/scrpt.cpk) with the current English
translation (texts/en/scrpt.cpk) on a stable (file, id) key, and emits one
row per visible string with columns:

  route | chapter | file | id | type | speaker | japanese | english | proposed_fix | notes

You browse/search this in Excel (or Excel for the web via OneDrive) while
playing. When a line is wrong, type the corrected text in `proposed_fix` and
optionally a tag in `notes` (e.g. "He->She"). Then run apply_review_fixes.py
to write only the filled-in rows back into texts/en/scrpt.cpk/<file>.csv.

The `file`+`id` columns are the round-trip key -- don't edit them. They're
greyed to signal that; `proposed_fix`/`notes` are tinted to invite input.

Regenerable: re-run any time the translation changes. Existing proposed_fix
edits live in the .xlsx (kept out of git) until applied, so regenerating
overwrites them -- apply first, or back up the sheet, before rebuilding.

Requires: pip install openpyxl
"""
import csv
import glob
import json
import os
import re
import sys

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EN_DIR = os.path.join(ROOT, "texts", "en", "scrpt.cpk")
JA_DIR = os.path.join(ROOT, "texts", "ja", "scrpt.cpk")
EN_ROOT = os.path.join(ROOT, "texts", "en")
JA_ROOT = os.path.join(ROOT, "texts", "ja")
SPEAKER_MAP = os.path.join(ROOT, "texts", "en", "_speaker_names.json")
OUT_XLSX = os.path.join(ROOT, "issues", "translation_review.xlsx")

COLUMNS = ["route", "chapter", "file", "id", "type",
           "speaker", "speaker_ja", "japanese", "english", "proposed_fix", "notes"]

# Non-script text files (UI / tips / etc.), each with a texts/ja/<name>.csv source.
# (stem, route label)
AUX_FILES = [
    ("AppGameDataTipsData", "(tips)"),
    ("FlowChartData", "(flowchart)"),
    ("Metadata", "(metadata)"),
    ("Text", "(ui)"),
    ("TextFlyMoveData", "(flymove)"),
]

# id prefix -> friendly row type
TYPE_MAP = {
    "XMESS": "dialogue",
    "XCHAPTITLE": "chapter title",
    "XCOMMENTVIEW": "caption",
    "MSTD": "system",
    "XRETURNTODEATH": "system",
    "tip_title": "tip title",
    "tip_explanation": "tip text",
    "flow_title": "flow title",
    "flow_information": "flow info",
    "metadata": "metadata",
    "text": "ui text",
    "fly": "fly text",
}

FNAME_RE = re.compile(r"^ST_(.+)_(\d+)([A-Za-z]*)$")


def read_csv(path):
    with open(path, encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def row_type(rid):
    rid = rid or ""
    if "#" in rid:                       # aux ids: tip_title#000, flow_title#00-001
        prefix = rid.split("#", 1)[0]
    else:                                # script ids: XMESS_0000
        m = re.match(r"([A-Za-z]+)", rid)
        prefix = m.group(1) if m else ""
    return TYPE_MAP.get(prefix, prefix.lower())


def parse_fname(stem):
    """ST_N_HDR_007A -> ('N_HDR', 7, 'A'); fallback keeps it sortable."""
    m = FNAME_RE.match(stem)
    if not m:
        return (stem, 0, "")
    return (m.group(1), int(m.group(2)), m.group(3))


def clean(s):
    return (s or "").strip()


def chapter_title(rows):
    for r in rows:
        if (r.get("id") or "").startswith("XCHAPTITLE"):
            return " ".join(clean(r.get("target")).split())
    return ""


def main():
    speakers = {}
    if os.path.exists(SPEAKER_MAP):
        with open(SPEAKER_MAP, encoding="utf-8") as f:
            speakers = json.load(f)

    records = []
    for en_path in sorted(glob.glob(os.path.join(EN_DIR, "*.csv"))):
        stem = os.path.splitext(os.path.basename(en_path))[0]
        route, num, sub = parse_fname(stem)
        en_rows = read_csv(en_path)
        chapter = chapter_title(en_rows)

        ja_path = os.path.join(JA_DIR, os.path.basename(en_path))
        ja_by_id = {}
        if os.path.exists(ja_path):
            for r in read_csv(ja_path):
                ja_by_id.setdefault(r.get("id"), r.get("target") or "")

        for idx, r in enumerate(en_rows):
            rid = r.get("id") or ""
            en_text = r.get("target") or ""
            ja_text = ja_by_id.get(rid, "")
            if not clean(en_text) and not clean(ja_text):
                continue
            dev = clean(r.get("developer_comments"))
            speaker = speakers.get(dev, dev)
            records.append({
                "sort": (0, route, num, sub, idx),
                "route": route,
                "chapter": chapter,
                "file": stem,
                "id": rid,
                "type": row_type(rid),
                "speaker": speaker,
                "speaker_ja": dev,
                "japanese": ja_text,
                "english": en_text,
                "proposed_fix": "",
                "notes": "",
            })

    # --- non-script text files (UI / tips / flowchart / metadata / flymove) ---
    for ai, (stem, label) in enumerate(AUX_FILES):
        en_path = os.path.join(EN_ROOT, stem + ".csv")
        if not os.path.exists(en_path):
            continue
        ja_path = os.path.join(JA_ROOT, stem + ".csv")
        ja_by_id = {}
        if os.path.exists(ja_path):
            for r in read_csv(ja_path):
                ja_by_id.setdefault(r.get("id"), r.get("target") or "")
        for idx, r in enumerate(read_csv(en_path)):
            rid = r.get("id") or ""
            en_text = r.get("target") or ""
            ja_text = ja_by_id.get(rid, "")
            if not clean(en_text) and not clean(ja_text):
                continue
            records.append({
                "sort": (1, label, ai, "", idx),
                "route": label,
                "chapter": "",
                "file": stem,
                "id": rid,
                "type": row_type(rid),
                "speaker": "",
                "speaker_ja": "",
                "japanese": ja_text,
                "english": en_text,
                "proposed_fix": "",
                "notes": "",
            })

    records.sort(key=lambda x: x["sort"])
    write_xlsx(records)
    print(f"rows: {len(records)}")
    print(f"wrote: {OUT_XLSX}")


def write_xlsx(records):
    wb = Workbook()
    ws = wb.active
    ws.title = "Dialogue"

    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF")
    ref_fill = PatternFill("solid", fgColor="F2F2F2")       # greyed reference cols
    key_fill = PatternFill("solid", fgColor="E2E2E2")       # greyed round-trip key
    edit_fill = PatternFill("solid", fgColor="FFF7DC")      # tinted editable cols
    top = Alignment(vertical="top", wrap_text=True)
    top_nowrap = Alignment(vertical="top", wrap_text=False)

    # header
    for c, name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")
    ws.cell(row=1, column=COLUMNS.index("file") + 1).comment = Comment(
        "Round-trip key - do not edit (file & id locate the line for apply_review_fixes.py).",
        "build_review_sheet")
    ws.cell(row=1, column=COLUMNS.index("proposed_fix") + 1).comment = Comment(
        "Type the corrected English here. Leave blank to keep the current line.",
        "build_review_sheet")

    wrap_cols = {"chapter", "japanese", "english", "proposed_fix", "notes"}
    key_cols = {"file", "id"}
    edit_cols = {"proposed_fix", "notes"}

    for rec in records:
        row = [rec[c] for c in COLUMNS]
        ws.append(row)
        r = ws.max_row
        for c, name in enumerate(COLUMNS, 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = top if name in wrap_cols else top_nowrap
            if name in edit_cols:
                cell.fill = edit_fill
            elif name in key_cols:
                cell.fill = key_fill
            else:
                cell.fill = ref_fill

    widths = {"route": 10, "chapter": 26, "file": 16, "id": 16, "type": 11,
              "speaker": 20, "speaker_ja": 16, "japanese": 42, "english": 42,
              "proposed_fix": 42, "notes": 22}
    for c, name in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(c)].width = widths[name]

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{ws.max_row}"

    os.makedirs(os.path.dirname(OUT_XLSX), exist_ok=True)
    wb.save(OUT_XLSX)


if __name__ == "__main__":
    sys.exit(main())
