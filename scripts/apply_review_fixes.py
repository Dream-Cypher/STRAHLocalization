#!/usr/bin/env python3
"""Apply proposed_fix edits from the review spreadsheet back into the EN CSVs.

Reads issues/translation_review.xlsx, and for every row whose `proposed_fix`
is non-empty and differs from the current `english`, writes that text into
texts/en/scrpt.cpk/<file>.csv at the matching `id`.

Dry-run by default (shows what would change). Pass --apply to write.
A change log is written to issues/applied_fixes_<timestamp>.csv either way.

Round-trips on the (file, id) key, so it's robust to any row-count / ordering
differences between JA and EN. The committed CSVs are the source of truth;
git is your undo (review `git diff texts/en` before committing).

Requires: pip install openpyxl
"""
import csv
import datetime as dt
import os
import sys

from openpyxl import load_workbook

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EN_DIR = os.path.join(ROOT, "texts", "en", "scrpt.cpk")   # dialogue scripts
EN_ROOT = os.path.join(ROOT, "texts", "en")               # aux files (tips/ui/...)
XLSX = os.path.join(ROOT, "issues", "translation_review.xlsx")


def resolve_csv(fname):
    """Map a sheet `file` value to its CSV path (script dir, else aux root)."""
    p1 = os.path.join(EN_DIR, fname + ".csv")
    if os.path.exists(p1):
        return p1
    p2 = os.path.join(EN_ROOT, fname + ".csv")
    if os.path.exists(p2):
        return p2
    return None


def read_sheet():
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb["Dialogue"] if "Dialogue" in wb.sheetnames else wb.active
    rows = ws.iter_rows(values_only=True)
    header = [str(h) if h is not None else "" for h in next(rows)]
    idx = {name: i for i, name in enumerate(header)}
    for need in ("file", "id", "english", "proposed_fix"):
        if need not in idx:
            sys.exit(f"ERROR: sheet missing '{need}' column (got {header})")
    out = []
    for r in rows:
        get = lambda n: (r[idx[n]] if idx[n] < len(r) else None)
        fix = get("proposed_fix")
        fix = "" if fix is None else str(fix)
        if not fix.strip():
            continue
        eng = get("english")
        eng = "" if eng is None else str(eng)
        if fix == eng:
            continue
        out.append({
            "file": str(get("file") or ""),
            "id": str(get("id") or ""),
            "speaker": str(get("speaker") or ""),
            "old": eng,
            "new": fix,
        })
    return out


def read_csv(path):
    with open(path, encoding="utf-8-sig", newline="") as f:
        rd = csv.DictReader(f)
        return rd.fieldnames, list(rd)


def write_csv(path, fieldnames, rows):
    # newline="" + \r\n keeps Excel/Git-friendly CRLF, matching the repo's CSVs
    with open(path, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames, lineterminator="\r\n")
        w.writeheader()
        w.writerows(rows)


def main():
    apply = "--apply" in sys.argv[1:]
    if not os.path.exists(XLSX):
        sys.exit(f"ERROR: {XLSX} not found -- run build_review_sheet.py first.")

    fixes = read_sheet()
    if not fixes:
        print("No proposed_fix entries found. Nothing to do.")
        return

    by_file = {}
    for fx in fixes:
        by_file.setdefault(fx["file"], []).append(fx)

    applied, skipped = [], []
    for fname, items in sorted(by_file.items()):
        path = resolve_csv(fname)
        if path is None:
            for fx in items:
                skipped.append((fx, "file not found"))
            continue
        fieldnames, rows = read_csv(path)
        by_id = {}
        for row in rows:
            by_id.setdefault(row.get("id"), row)
        changed = False
        for fx in items:
            row = by_id.get(fx["id"])
            if row is None:
                skipped.append((fx, "id not found"))
                continue
            if (row.get("target") or "") == fx["new"]:
                continue  # already applied
            row["target"] = fx["new"]
            changed = True
            applied.append(fx)
        if changed and apply:
            write_csv(path, fieldnames, rows)

    ts = dt.datetime.now().strftime("%Y%m%d-%H%M%S")
    log = os.path.join(ROOT, "issues", f"applied_fixes_{ts}.csv")
    os.makedirs(os.path.dirname(log), exist_ok=True)
    with open(log, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["file", "id", "speaker", "old", "new", "status"])
        for fx in applied:
            w.writerow([fx["file"], fx["id"], fx["speaker"], fx["old"], fx["new"],
                        "applied" if apply else "would-apply"])
        for fx, why in skipped:
            w.writerow([fx["file"], fx["id"], fx["speaker"], fx["old"], fx["new"], f"SKIP: {why}"])

    verb = "Applied" if apply else "Would apply"
    print(f"{verb} {len(applied)} fix(es) across {len(by_file)} file(s).")
    if skipped:
        print(f"Skipped {len(skipped)} (see log).")
    print(f"Log: {log}")
    if not apply:
        print("\nDRY RUN -- no files changed. Re-run with --apply to write, "
              "then review `git diff texts/en` before committing.")


if __name__ == "__main__":
    sys.exit(main())
