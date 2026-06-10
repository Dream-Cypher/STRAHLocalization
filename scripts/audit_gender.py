#!/usr/bin/env python3
"""Flag likely he/she pronoun errors and seed the review sheet's `notes` column.

Japanese frequently omits pronouns, so the original machine translation often
guessed gender wrong. The Chinese localization (texts/zh_Hans) and, where
present, Japanese DO mark third-person gender (Chinese 他/她, Japanese 彼/彼女),
so we flag dialogue lines where the English pronoun gender DISAGREES with the
Chinese/Japanese reference. Conservative: a line is flagged only when English
is unambiguously one gender and the reference is unambiguously the other.

Writes a "GENDER(...)" tag into the `notes` column of
issues/translation_review.xlsx (idempotent: re-running refreshes its own tag
and preserves any other note text) and a report to issues/gender_audit_<ts>.csv,
which includes each flag's surrounding dialogue lines (speaker + +/-2 lines) so
the referent can be judged in context rather than from the line alone.

Confidence: he->she flags use Chinese 她 (reliably female) = high; she->he
flags lean on 他/彼 (more false-friends) = lower -- review those carefully.

With --prefill, the high-conf (he->she) rows also get a mechanical, case-
preserving pronoun swap written into `proposed_fix` (he->she, him/his->her,
himself->herself) as a starting point -- only when that cell is empty, so it
never clobbers your edits. Still verify each: some need a clause rephrase, not
just a swap. The low-conf (she->he) rows are never auto-filled.

Requires: pip install openpyxl   (run via the project .venv)
"""
import csv
import datetime as dt
import glob
import os
import re

from openpyxl import load_workbook

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EN = os.path.join(ROOT, "texts", "en", "scrpt.cpk")
JA = os.path.join(ROOT, "texts", "ja", "scrpt.cpk")
ZH = os.path.join(ROOT, "texts", "zh_Hans", "scrpt.cpk")
XLSX = os.path.join(ROOT, "issues", "translation_review.xlsx")
TAG = "GENDER"

RE_M = re.compile(r"\b(he|him|his|himself)\b", re.I)
RE_F = re.compile(r"\b(she|her|hers|herself)\b", re.I)

# mechanical male->female pronoun swap for high-conf prefill (case-preserving).
# his->her covers the common determiner ("his car"->"her car"); rare predicative
# "hers" and any clause rephrasing are left for the human reviewer.
_M2F = {"he": "she", "him": "her", "his": "her", "himself": "herself"}


def swap_m2f(text):
    def repl(m):
        w = m.group(0)
        r = _M2F[w.lower()]
        if w.isupper():
            return r.upper()
        if w[0].isupper():
            return r[0].upper() + r[1:]
        return r
    return RE_M.sub(repl, text)
# Chinese 他 false-friends to strip before counting it as a male pronoun
ZH_MALE_NOISE = ["其他", "吉他", "他人", "他乡", "他日", "他方", "他用"]


def read(path):
    if not os.path.exists(path):
        return {}
    with open(path, encoding="utf-8-sig", newline="") as f:
        return {r["id"]: (r["target"] or "") for r in csv.DictReader(f)}


def read_rows(path):
    if not os.path.exists(path):
        return []
    with open(path, encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def load_speakers():
    p = os.path.join(ROOT, "texts", "en", "_speaker_names.json")
    if os.path.exists(p):
        import json
        with open(p, encoding="utf-8") as f:
            return json.load(f)
    return {}


def build_context(dlg, i, speakers, span=2):
    """Lines around dlg[i] as 'Speaker: line' (>> marks the flagged line)."""
    out = []
    for j in range(max(0, i - span), min(len(dlg), i + span + 1)):
        rid, txt, dev = dlg[j]
        name = speakers.get(dev, dev) or "?"
        one = " ".join((txt or "").split())
        out.append(f"{'>> ' if j == i else '   '}{name}: {one}")
    return "\n".join(out)


def en_words(t):
    m = sorted({w.lower() for w in RE_M.findall(t)})
    f = sorted({w.lower() for w in RE_F.findall(t)})
    return m, f


def zh_gender(t):
    female = "她" in t
    s = t
    for n in ZH_MALE_NOISE:
        s = s.replace(n, "")
    male = "他" in s
    return male, female


def ja_gender(t):
    nf = t.count("彼女")
    return (t.count("彼") - nf) > 0, nf > 0


def find_candidates():
    speakers = load_speakers()
    cands = []
    for enp in sorted(glob.glob(os.path.join(EN, "*.csv"))):
        b = os.path.basename(enp)
        ja = read(os.path.join(JA, b))
        zh = read(os.path.join(ZH, b))
        # ordered dialogue lines, so we can pull surrounding context
        dlg = [(r["id"], r["target"] or "", r["developer_comments"] or "")
               for r in read_rows(enp)
               if r["id"].startswith("XMESS") and (r["target"] or "").strip()]
        for i, (rid, et, dev) in enumerate(dlg):
            em_w, ef_w = en_words(et)
            em, ef = bool(em_w), bool(ef_w)
            if not (em ^ ef):          # need exactly one EN gender
                continue
            zt, jt = zh.get(rid, ""), ja.get(rid, "")
            zm, zf = zh_gender(zt)
            jm, jf = ja_gender(jt)
            ref_f, ref_m = (zf or jf), (zm or jm)
            refs = []
            spk = speakers.get(dev, dev)
            ctx = build_context(dlg, i, speakers)
            if em and not ef and ref_f and not ref_m:
                if zf: refs.append("ZH=她")
                if jf: refs.append("JA=彼女")
                cands.append(dict(file=b[:-4], id=rid, dir="he->she", conf="high",
                                  en="/".join(em_w), refs=" ".join(refs),
                                  suggest="she/her", en_text=et, zh=zt, ja=jt,
                                  speaker=spk, context=ctx))
            elif ef and not em and ref_m and not ref_f:
                if zm: refs.append("ZH=他")
                if jm: refs.append("JA=彼")
                cands.append(dict(file=b[:-4], id=rid, dir="she->he", conf="low",
                                  en="/".join(ef_w), refs=" ".join(refs),
                                  suggest="he/him", en_text=et, zh=zt, ja=jt,
                                  speaker=spk, context=ctx))
    return cands


def note_for(c):
    return f"{TAG}({c['conf']}): EN={c['en']} vs {c['refs']} -> check {c['suggest']}"


def seed_sheet(cands, prefill=False):
    if not os.path.exists(XLSX):
        print(f"(workbook not found: {XLSX} -- skipping note seeding; report still written)")
        return 0, 0
    wb = load_workbook(XLSX)
    ws = wb["Dialogue"] if "Dialogue" in wb.sheetnames else wb.active
    hdr = [c.value for c in ws[1]]
    ci = {n: i for i, n in enumerate(hdr)}
    fcol, icol, ncol, pcol, ecol = (ci["file"], ci["id"], ci["notes"],
                                    ci["proposed_fix"], ci["english"])
    by_key = {(c["file"], c["id"]): c for c in cands}
    seeded = prefilled = 0
    for row in ws.iter_rows(min_row=2):
        key = (row[fcol].value, row[icol].value)
        c = by_key.get(key)
        cell = row[ncol]
        existing = cell.value or ""
        # strip any prior GENDER tag so re-runs don't duplicate
        kept = " | ".join(p for p in str(existing).split(" | ")
                          if p and not p.startswith(TAG + "("))
        if c:
            cell.value = note_for(c) + (" | " + kept if kept else "")
            seeded += 1
            # high-conf (he->she) only: seed a swap, never clobber an existing edit
            if prefill and c["conf"] == "high" and not (row[pcol].value or "").strip():
                eng = row[ecol].value or ""
                swapped = swap_m2f(eng)
                if swapped != eng:
                    row[pcol].value = swapped
                    prefilled += 1
        elif kept != existing:
            cell.value = kept
    wb.save(XLSX)
    return seeded, prefilled


def write_report(cands):
    ts = dt.datetime.now().strftime("%Y%m%d-%H%M%S")
    path = os.path.join(ROOT, "issues", f"gender_audit_{ts}.csv")
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["file", "id", "direction", "confidence", "speaker", "en_pronoun",
                    "reference", "english", "chinese", "japanese", "context"])
        for c in cands:
            w.writerow([c["file"], c["id"], c["dir"], c["conf"], c["speaker"],
                        c["en"], c["refs"], c["en_text"], c["zh"], c["ja"],
                        c["context"]])
    return path


def main():
    import sys
    prefill = "--prefill" in sys.argv[1:]
    cands = find_candidates()
    hi = sum(1 for c in cands if c["conf"] == "high")
    lo = len(cands) - hi
    seeded, prefilled = seed_sheet(cands, prefill=prefill)
    report = write_report(cands)
    print(f"candidates: {len(cands)}  (he->she high-conf: {hi}, she->he low-conf: {lo})")
    print(f"seeded notes in workbook: {seeded}")
    if prefill:
        print(f"pre-filled proposed_fix (high-conf, mechanical swap): {prefilled}")
    else:
        print("(notes only; pass --prefill to seed high-conf proposed_fix swaps)")
    print(f"report: {report}")


if __name__ == "__main__":
    main()
